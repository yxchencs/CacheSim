import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def save_dataframe_to_excel_with_format(df, file_name):
    # 创建一个 ExcelWriter 对象
    excel_writer = pd.ExcelWriter(file_name, engine='openpyxl')
    df.to_excel(excel_writer, index=False, sheet_name='Sheet1')

    # 获取当前活动的工作表
    workbook = excel_writer.book
    worksheet = workbook.active

    # 设置字体样式（不加粗）
    font = Font(name='Arial', size=12, bold=False)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.font = font

    # 居中单元格内容
    alignment = Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.alignment = alignment

    # 调整单元格宽度
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    # 保存 Excel 文件
    excel_writer.save()
    excel_writer.close()


def str_to_datetime(raw_str, format):
    return datetime.strptime(raw_str, format)


def extract_statistic(filename, rdwr_only):
    with open(filename, 'r') as file:
        data = file.read()
    bias = 0
    if not rdwr_only:
        bias = 2
    lines = data.strip().split('\n')
    # print(lines[13])
    policy = lines[0].split()[2]
    trace_hit_ratio = float(lines[13 + bias].split()[3])
    # power = float(lines[21+bias].split()[1])
    total_time = float(lines[15 + bias].split()[2])
    p99 = float(lines[17 + bias].split('=')[2].split()[0])
    avg_latency = float(lines[16 + bias].split()[2])
    bandwidth = float(lines[20 + bias].split()[1])

    emmc_read_nums = int(lines[22 + bias].split(';')[0].strip().split()[2])
    emmc_read_avg_latency = float(lines[22 + bias].split(';')[1].strip().split()[2])
    emmc_read_p99 = float(lines[22 + bias].split(';')[2].strip().split()[8])

    emmc_write_nums = int(lines[23 + bias].split(';')[0].strip().split()[2])
    emmc_write_avg_latency = float(lines[23 + bias].split(';')[1].strip().split()[2])
    emmc_write_p99 = float(lines[23 + bias].split(';')[2].strip().split()[8])
    
    sd_read_nums = int(lines[24 + bias].split(';')[0].strip().split()[2])
    sd_read_avg_latency = float(lines[24 + bias].split(';')[1].strip().split()[2])
    sd_read_p99 = float(lines[24 + bias].split(';')[2].strip().split()[8])

    sd_write_nums = int(lines[25 + bias].split(';')[0].strip().split()[2])
    sd_write_avg_latency = float(lines[25 + bias].split(';')[1].strip().split()[2])
    sd_write_p99 = float(lines[25 + bias].split(';')[2].strip().split()[8])


    time_format = "%Y/%m/%d %H:%M:%S"

    time_begin = lines[14 + bias].split('to')[0].split('From')[1].strip()
    time_begin = str_to_datetime(time_begin, time_format)

    time_end = lines[14 + bias].split('to')[1].strip()
    time_end = str_to_datetime(time_end, time_format)

    # print(time_begin, time_end)

    return policy, trace_hit_ratio, total_time,\
            p99, avg_latency, time_begin, time_end, bandwidth, \
            emmc_read_nums, emmc_read_avg_latency, emmc_read_p99, \
            emmc_write_nums, emmc_write_avg_latency, emmc_write_p99, \
            sd_read_nums, sd_read_avg_latency, sd_read_p99, \
            sd_write_nums, sd_write_avg_latency, sd_write_p99 \


def extract_statistic_no_cache(filename, rdwr_only):
    with open(filename, 'r') as file:
        data = file.read()
    bias = 0
    if not rdwr_only:
        bias = 2
    lines = data.strip().split('\n')
    total_time = float(lines[15 + bias].split()[2])
    p99 = float(lines[17 + bias].split('=')[2].split()[0])
    avg_latency = float(lines[16 + bias].split()[2])
    bandwidth = float(lines[20 + bias].split()[1])

    time_format = "%Y/%m/%d %H:%M:%S"

    time_begin = lines[14 + bias].split('to')[0].split('From')[1].strip()
    time_begin = str_to_datetime(time_begin, time_format)

    time_end = lines[14 + bias].split('to')[1].strip()
    time_end = str_to_datetime(time_end, time_format)

    return total_time, p99, avg_latency, time_begin, time_end, bandwidth


def extract_statistic_except_power(filename):
    with open(filename, 'r') as file:
        data = file.read()

    lines = data.strip().split('\n')
    # print(lines[12])
    trace_hit_ratio = float(lines[15].split()[3])
    # power = float(lines[20].split()[1])
    total_time = float(lines[17].split()[2])
    p99 = float(lines[19].split('=')[2].split()[0])
    avg_latency = float(lines[18].split()[2])

    time_format = "%Y/%m/%d %H:%M:%S"

    time_begin = lines[16].split('to')[0].split('From')[1].strip()
    time_begin = str_to_datetime(time_begin, time_format)

    time_end = lines[16].split('to')[1].strip()
    time_end = str_to_datetime(time_end, time_format)

    # print(time_begin, time_end)

    return trace_hit_ratio, total_time, p99, avg_latency, time_begin, time_end


def extract_statistic_device_test(filename, rdwr_only):
    with open(filename, 'r') as file:
        data = file.read()
    bias = 0
    if not rdwr_only:
        bias = 2
    lines = data.strip().split('\n')
    total_time = float(lines[15 + bias].split()[2])
    avg_latency = float(lines[16 + bias].split()[2])
    p99 = float(lines[17 + bias].split('=')[2].split()[0])
    bandwidth = float(lines[20 + bias].split()[1])

    emmc_read_nums = int(lines[22 + bias].split(';')[0].strip().split()[2])
    emmc_read_avg_latency = float(lines[22 + bias].split(';')[1].strip().split()[2])
    emmc_read_p99 = float(lines[22 + bias].split(';')[2].strip().split()[8])

    emmc_write_nums = int(lines[23 + bias].split(';')[0].strip().split()[2])
    emmc_write_avg_latency = float(lines[23 + bias].split(';')[1].strip().split()[2])
    emmc_write_p99 = float(lines[23 + bias].split(';')[2].strip().split()[8])
    
    sd_read_nums = int(lines[24 + bias].split(';')[0].strip().split()[2])
    sd_read_avg_latency = float(lines[24 + bias].split(';')[1].strip().split()[2])
    sd_read_p99 = float(lines[24 + bias].split(';')[2].strip().split()[8])

    sd_write_nums = int(lines[25 + bias].split(';')[0].strip().split()[2])
    sd_write_avg_latency = float(lines[25 + bias].split(';')[1].strip().split()[2])
    sd_write_p99 = float(lines[25 + bias].split(';')[2].strip().split()[8])


    time_format = "%Y/%m/%d %H:%M:%S"

    time_begin = lines[14 + bias].split('to')[0].split('From')[1].strip()
    time_begin = str_to_datetime(time_begin, time_format)

    time_end = lines[14 + bias].split('to')[1].strip()
    time_end = str_to_datetime(time_end, time_format)

    # print(time_begin, time_end)

    return total_time,\
            p99, avg_latency, time_begin, time_end, bandwidth, \
            emmc_read_nums, emmc_read_avg_latency, emmc_read_p99, \
            emmc_write_nums, emmc_write_avg_latency, emmc_write_p99, \
            sd_read_nums, sd_read_avg_latency, sd_read_p99, \
            sd_write_nums, sd_write_avg_latency, sd_write_p99 \


def calculate_avg_cpu_usage(file_path, time_begin, time_end):
    cpu_cnt, cpu_sum = 0, 0
    time_format = "%Y-%m-%d %H:%M:%S"

    with open(file_path, 'r') as file:
        for line in file:
            # print(line)
            parts = line.strip().split()
            time = str_to_datetime((parts[1] + ' ' + parts[2]), time_format)

            if time_end >= time >= time_begin:
                cpu_cnt += 1
                cpu_sum += float(parts[0].split('%')[0].strip())

    if cpu_cnt != 0:
        return cpu_sum / cpu_cnt
    else:
        return None


def calculate_avg_power(file_path, time_begin, time_end):
    power_cnt, power_sum = 0, 0
    # Assuming the date is not important for the time in the Excel file, and only time is relevant
    time_format = "%H:%M:%S"
    # Convert string parameters to time objects
    # print('calculate_avg_power->', str(time_begin).split()[-1])
    time_begin_obj = datetime.strptime(str(time_begin).split()[-1], time_format)
    time_end_obj = datetime.strptime(str(time_end).split()[-1], time_format)
    # print(time_begin_obj, time_end_obj)
    # Read the Excel file
    df = pd.read_excel(file_path)

    # Calculate the average power between time_begin and time_end
    for index, row in df.iterrows():
        # Extract the time from the Excel file
        # print(row['时间'])
        if str(row['时间']) != '时间':
            current_time_obj = datetime.strptime(str(row['时间']), time_format)
            # Check if the current time is within the time range specified
            if time_end_obj >= current_time_obj >= time_begin_obj:
                power_cnt += 1
                power_sum += float(row['功率'])
    # print("power_cnt:",power_cnt)
    if power_cnt == 0:
        return None
    return power_sum / power_cnt


# [time_begin, time_end]时间段所有mem_used的平均值
def calculate_avg_mem_used(file_path, time_begin, time_end):
    mem_begin, mem_cnt, mem_sum = 0, 0, 0
    time_format = "%Y-%m-%d %H:%M:%S"
    # with open(file_path, 'r') as file:
    #     for line in file:
    #         parts = line.strip().split()
    #         time = str_to_datetime((parts[1] + ' ' + parts[2]), time_format)
    #
    #         if time >= time_begin:
    #             mem_begin = int(parts[0].split('MB')[0])
    #             break

    with open(file_path, 'r') as file:
        for line in file:
            parts = line.strip().split()
            time = str_to_datetime((parts[1] + ' ' + parts[2]), time_format)

            if time_end >= time >= time_begin:
                mem_used = int(parts[0].split('MB')[0].strip())
                if mem_begin == 0:
                    mem_begin = mem_used
                mem_cnt += 1
                # mem_sum += int(parts[0].split('MB')[0].strip()) - mem_begin # [time_begin, time_end]时间段内所有mem_used数值和mem_begin的差之和 / mem_cnt
                mem_sum += mem_used

    if mem_cnt != 0:
        return mem_sum / mem_cnt - mem_begin
    else:
        return None

# IO Read/Write Mount
def calculate_disk_read_wrtn(file_path, time_begin, time_end):
    emmc_kb_read_begin, emmc_kb_wrtn_begin, sd_kb_read_begin, sd_kb_wrtn_begin = 0, 0, 0, 0
    emmc_kb_read_end, emmc_kb_wrtn_end, sd_kb_read_end, sd_kb_wrtn_end = 0, 0, 0, 0
    time_format = "%Y-%m-%d %H:%M:%S"

    with open(file_path, 'r') as file:
        for line in file:
            # print(line)
            parts = line.strip().split(';')
            time = str_to_datetime((parts[2].strip()), time_format)

            if time == time_begin:
                emmc = parts[0].split()
                sd = parts[1].split()
                emmc_kb_read_begin = int(emmc[3].strip(','))
                emmc_kb_wrtn_begin = int(emmc[6].strip())
                sd_kb_read_begin = int(sd[3].strip(','))
                sd_kb_wrtn_begin = int(sd[6].strip())
                break
            if time == time_begin + timedelta(seconds=1):
                emmc = parts[0].split()
                sd = parts[1].split()
                emmc_kb_read_begin = int(emmc[3].strip(','))
                emmc_kb_wrtn_begin = int(emmc[6].strip())
                sd_kb_read_begin = int(sd[3].strip(','))
                sd_kb_wrtn_begin = int(sd[6].strip())
                break

    with open(file_path, 'r') as file:
        for line in file:
            # print(line)
            parts = line.strip().split(';')
            time = str_to_datetime((parts[2].strip()), time_format)

            if time == time_end:
                emmc = parts[0].split()
                sd = parts[1].split()
                emmc_kb_read_end = int(emmc[3].strip(','))
                emmc_kb_wrtn_end = int(emmc[6].strip())
                sd_kb_read_end = int(sd[3].strip(','))
                sd_kb_wrtn_end = int(sd[6].strip())
                break
            if time == time_end + timedelta(seconds=1):
                emmc = parts[0].split()
                sd = parts[1].split()
                emmc_kb_read_end = int(emmc[3].strip(','))
                emmc_kb_wrtn_end = int(emmc[6].strip())
                sd_kb_read_end = int(sd[3].strip(','))
                sd_kb_wrtn_end = int(sd[6].strip())
                break

    emmc_kb_read = emmc_kb_read_end - emmc_kb_read_begin
    emmc_kb_wrtn = emmc_kb_wrtn_end - emmc_kb_wrtn_begin
    sd_kb_read = sd_kb_read_end - sd_kb_read_begin
    sd_kb_wrtn = sd_kb_wrtn_end - sd_kb_wrtn_begin
    return emmc_kb_read, emmc_kb_wrtn, sd_kb_read, sd_kb_wrtn


def process_results_no_cache():
    trace_type_list = ['latest', 'uniform', 'zipfian']
    operation_read_ratio_list = ['read_0', 'read_1']
    device_list = ['emmc', 'sd']

    folder_list = []  # 保存所有遍历过的trace目录

    list_trace_list, list_operation_read_ratio, list_device, list_trace_hit_ratio, list_total, list_p99, list_avg_latency, list_cpu_usage, list_mem_used, list_emmc_kb_read, \
    list_emmc_kb_wrtn, list_sd_kb_read, list_sd_kb_wrtn, list_avg_power = [], [], [], [], [], [], [], [], [], [], [], [], [], []
    list_time_begin, list_time_end = [], []
    list_bandwidth = []

    for type in trace_type_list:
        for ratio in operation_read_ratio_list:
            rdwr_only = False
            if ratio == 'read_1' or ratio == 'read_0':
                rdwr_only = True
            for device in device_list:
                list_trace_list.append(type)
                list_operation_read_ratio.append(ratio)
                list_device.append(device)
                file_path_begin = path_head + type + '/' + ratio + '/' + device + '/'
                print('data process:', file_path_begin)
                file_statistic_path = file_path_begin + file_statistic_name

                total, p99, avg_latency, time_begin, time_end, bandwidth = extract_statistic_no_cache(
                    file_statistic_path,
                    rdwr_only)
                list_cpu_usage.append(
                    calculate_avg_cpu_usage(file_cpu_usage_path, time_begin, time_end - timedelta(seconds=1)))
                list_mem_used.append(calculate_avg_mem_used(file_mem_used_path, time_begin, time_end))
                emmc_kb_read, emmc_kb_wrtn, sd_kb_read, sd_kb_wrtn \
                    = calculate_disk_read_wrtn(file_disk_path, time_begin, time_end)
                list_avg_power.append(calculate_avg_power(file_power_path, time_begin, time_end))

                list_emmc_kb_read.append(emmc_kb_read)
                list_emmc_kb_wrtn.append(emmc_kb_wrtn)
                list_sd_kb_read.append(sd_kb_read)
                list_sd_kb_wrtn.append(sd_kb_wrtn)

                # list_random.append(random)

                list_total.append(total)
                list_p99.append(p99)
                list_avg_latency.append(avg_latency)

                list_time_begin.append(time_begin)
                list_time_end.append(time_end)
                # print(trace_hit_ratio, "\t", power, "\t\t", total, "\t", p99, "\t", avg_latency)

                list_bandwidth.append(bandwidth)

    # 计算 energy
    # 使用列表推导式计算两个列表中元素的分别相乘
    list_energy = [a * b for a, b in zip(list_avg_power, list_total)]

    # 将多个列表放入一个字典
    data = {'trace type': list_trace_list, 'operation read ratio': list_operation_read_ratio, 'device': list_device,
            'average latency/ms': list_avg_latency, 'P99 latency/ms': list_p99, 'average power/W': list_avg_power,
            'energy/J': list_energy, 'cpu usage/%': list_cpu_usage, 'memory used/MB': list_mem_used,
            'total time/s': list_total, "bandwidth/MB/s": list_bandwidth,
            'avg_emmc_read/KB': list_emmc_kb_read, 'avg_emmc_wrtn/KB': list_emmc_kb_wrtn,
            'avg_sd_read/KB': list_sd_kb_read, 'avg_sd_wrtn/KB': list_sd_kb_wrtn}

    # print(data)

    # 将字典转换为DataFrame
    df = pd.DataFrame(data)

    # 指定要保存的Excel文件名
    excel_file = path_head + 'statistic.xlsx'

    # save_dataframe_to_excel_with_format(df,excel_file)
    # 将DataFrame保存为Excel文件
    df.to_excel(excel_file, index=False)

    print(f"data save in {excel_file}")


def process_results():
    list_io, list_cache_size, list_policy = [], [], []
    list_trace_list, list_operation_read_ratio, list_device, list_trace_hit_ratio, list_total, list_p99, list_avg_latency = [], [], [], [], [], [], []
    list_cpu_usage, list_mem_used, list_emmc_kb_read, list_emmc_kb_wrtn, list_sd_kb_read, list_sd_kb_wrtn, list_avg_power = [], [], [], [], [], [], []
    list_time_begin, list_time_end = [], []
    list_bandwidth = []
    list_emmc_read_nums, list_emmc_read_avg_latency, list_emmc_read_p99 = [], [], []
    list_emmc_write_nums, list_emmc_write_avg_latency, list_emmc_write_p99 = [], [], []
    list_sd_read_nums, list_sd_read_avg_latency, list_sd_read_p99 = [], [], []
    list_sd_write_nums, list_sd_write_avg_latency, list_sd_write_p99 = [], [], []

    for type in trace_type_list:
        for ratio in operation_read_ratio_list:
            rdwr_only = False
            if ratio == 'read_1' or ratio == 'read_0':
                rdwr_only = True
            for io in io_list:
                for size in cache_size_list:
                    for policy in policy_list:
                        list_trace_list.append(type)
                        list_operation_read_ratio.append(ratio)
                        list_io.append(io)
                        list_cache_size.append(size)
                        list_policy.append(policy)
                        file_path_begin = os.path.join(path_head, type, ratio, io, size, policy)
                        print('data process:', file_path_begin)
                        file_statistic_path = os.path.join(file_path_begin, file_statistic_name)

                        policy_sta, hit_ratio, total, p99, avg_latency,\
                        time_begin, time_end, bandwidth, \
                        emmc_read_nums, emmc_read_avg_latency, emmc_read_p99, \
                        emmc_write_nums, emmc_write_avg_latency, emmc_write_p99, \
                        sd_read_nums, sd_read_avg_latency, sd_read_p99, \
                        sd_write_nums, sd_write_avg_latency, sd_write_p99 = \
                        extract_statistic(
                            file_statistic_path,
                            rdwr_only)
                        if policy_sta != policy:
                            print("Error! policy_sta != policy")
                            exit(1)

                        list_cpu_usage.append(
                            calculate_avg_cpu_usage(file_cpu_usage_path, time_begin,
                                                    time_end - timedelta(seconds=1)))
                        list_mem_used.append(calculate_avg_mem_used(file_mem_used_path, time_begin, time_end))
                        emmc_kb_read, emmc_kb_wrtn, sd_kb_read, sd_kb_wrtn \
                            = calculate_disk_read_wrtn(file_disk_path, time_begin, time_end)
                        list_avg_power.append(calculate_avg_power(file_power_path, time_begin, time_end))

                        list_emmc_kb_read.append(emmc_kb_read)
                        list_emmc_kb_wrtn.append(emmc_kb_wrtn)
                        list_sd_kb_read.append(sd_kb_read)
                        list_sd_kb_wrtn.append(sd_kb_wrtn)

                        # list_random.append(random)
                        list_trace_hit_ratio.append(hit_ratio)
                        list_total.append(total)
                        list_p99.append(p99)
                        list_avg_latency.append(avg_latency)

                        list_time_begin.append(time_begin)
                        list_time_end.append(time_end)
                        # print(trace_hit_ratio, "\t", power, "\t\t", total, "\t", p99, "\t", avg_latency)

                        list_bandwidth.append(bandwidth)

                        list_emmc_read_nums.append(emmc_read_nums)
                        list_emmc_read_avg_latency.append(emmc_read_avg_latency)
                        list_emmc_read_p99.append(emmc_read_p99)
                        list_emmc_write_nums.append(emmc_write_nums)
                        list_emmc_write_avg_latency.append(emmc_write_avg_latency)
                        list_emmc_write_p99.append(emmc_write_p99)
                        list_sd_read_nums.append(sd_read_nums)
                        list_sd_read_avg_latency.append(sd_read_avg_latency)
                        list_sd_read_p99.append(sd_read_p99)
                        list_sd_write_nums.append(sd_write_nums)
                        list_sd_write_avg_latency.append(sd_write_avg_latency)
                        list_sd_write_p99.append(sd_write_p99)

    list_energy = [a * b if a is not None and b is not None else None for a, b in zip(list_avg_power, list_total)]
    
    list_emmc_kb_read = [x / 1024 for x in list_emmc_kb_read]
    list_emmc_kb_wrtn = [x / 1024 for x in list_emmc_kb_wrtn]
    list_sd_kb_read = [x / 1024 for x in list_sd_kb_read]
    list_sd_kb_wrtn = [x / 1024 for x in list_sd_kb_wrtn]
    dict_cache_policy={'fifo':'FIFO','lfu':"LFU",'lru':'LRU',
                        'lirs':'LIRS','arc':'ARC','clockpro':'CLOCK-Pro',
                        'random':'Random','2q':'2Q','tinylfu':'TinyLFU'}
    list_policy = [dict_cache_policy[x] for x in list_policy]

    # 将多个列表放入一个字典
    data = {'Trace Pattern': list_trace_list, 'Operation Read Ratio': list_operation_read_ratio,
            'IO': list_io, 'Cache Size': list_cache_size, 'Cache Policy': list_policy,
            'Hit Ratio':list_trace_hit_ratio,
            'Average Latency(ms)': list_avg_latency, 'P99 Latency(ms)': list_p99, 
            'Average CPU Usage(%)': list_cpu_usage, 'Average Memory Used(MB)': list_mem_used,
            'Total Time(s)': list_total, "Bandwidth(MB/s)": list_bandwidth,
            'eMMC Read Mount(MB)': list_emmc_kb_read, 'eMMC Write Mount(MB)': list_emmc_kb_wrtn,
            'SD Read Mount(MB)': list_sd_kb_read, 'SD Write Mount(MB)': list_sd_kb_wrtn,
            'eMMC Read Numbers': list_emmc_read_nums, 'eMMC Read Avarage Latency(ms)': list_emmc_read_avg_latency, 'eMMC Read P99 Latency(ms)':list_emmc_read_p99,
            'eMMC Write Numbers': list_emmc_write_nums, 'eMMC Write Avarage Latency(ms)': list_emmc_write_avg_latency, 'eMMC Write P99 Latency(ms)':list_emmc_write_p99,
            'SD Read Numbers': list_sd_read_nums, 'SD Read Avarage Latency(ms)': list_sd_read_avg_latency, 'SD Read P99 Latency(ms)':list_sd_read_p99,
            'SD Write Numbers': list_sd_write_nums, 'SD Write Avarage Latency(ms)': list_sd_write_avg_latency, 'SD Write P99 Latency(ms)':list_sd_write_p99,
            'Average Power(W)': list_avg_power, 'Energy(J)': list_energy, }

    # print(data)

    # 将字典转换为DataFrame
    df = pd.DataFrame(data)

    # 指定要保存的Excel文件名
    excel_file = path_head + 'statistic.xlsx'

    # save_dataframe_to_excel_with_format(df,excel_file)
    # 将DataFrame保存为Excel文件
    df.to_excel(excel_file, index=False)

    print(f"data save in {excel_file}")

def process_results_except_power():
    list_io, list_cache_size, list_policy = [], [], []
    list_trace_list, list_operation_read_ratio, list_trace_hit_ratio, list_total, list_p99, list_avg_latency =  [], [], [], [], [], []
    list_cpu_usage, list_mem_used, list_emmc_kb_read, list_emmc_kb_wrtn, list_sd_kb_read, list_sd_kb_wrtn = [], [], [], [], [], []
    list_time_begin, list_time_end = [], []
    list_bandwidth = []
    list_emmc_read_nums, list_emmc_read_avg_latency, list_emmc_read_p99 = [], [], []
    list_emmc_write_nums, list_emmc_write_avg_latency, list_emmc_write_p99 = [], [], []
    list_sd_read_nums, list_sd_read_avg_latency, list_sd_read_p99 = [], [], []
    list_sd_write_nums, list_sd_write_avg_latency, list_sd_write_p99 = [], [], []

    for type in trace_type_list:
        for ratio in operation_read_ratio_list:
            rdwr_only = False
            if ratio == 'read_1' or ratio == 'read_0':
                rdwr_only = True
            for io in io_list:
                for size in cache_size_list:
                    for policy in policy_list:
                        list_trace_list.append(type)
                        list_operation_read_ratio.append(ratio)
                        list_io.append(io)
                        list_cache_size.append(size)
                        list_policy.append(policy)
                        file_path_begin = os.path.join(path_head, type, ratio, io, size, policy)
                        print('data process:', file_path_begin)
                        file_statistic_path = os.path.join(file_path_begin, file_statistic_name)

                        policy_sta, hit_ratio, total, p99, avg_latency, \
                        time_begin, time_end, bandwidth, \
                        emmc_read_nums, emmc_read_avg_latency, emmc_read_p99, \
                        emmc_write_nums, emmc_write_avg_latency, emmc_write_p99, \
                        sd_read_nums, sd_read_avg_latency, sd_read_p99, \
                        sd_write_nums, sd_write_avg_latency, sd_write_p99 = \
                        extract_statistic(
                            file_statistic_path,
                            rdwr_only)
                        
                        if policy_sta != policy:
                            print("Error! policy_sta != policy")
                            exit(1)

                        list_cpu_usage.append(
                            calculate_avg_cpu_usage(file_cpu_usage_path, time_begin,
                                                    time_end - timedelta(seconds=1)))
                        list_mem_used.append(calculate_avg_mem_used(file_mem_used_path, time_begin, time_end))
                        emmc_kb_read, emmc_kb_wrtn, sd_kb_read, sd_kb_wrtn \
                            = calculate_disk_read_wrtn(file_disk_path, time_begin, time_end)

                        list_emmc_kb_read.append(emmc_kb_read)
                        list_emmc_kb_wrtn.append(emmc_kb_wrtn)
                        list_sd_kb_read.append(sd_kb_read)
                        list_sd_kb_wrtn.append(sd_kb_wrtn)

                        list_trace_hit_ratio.append(hit_ratio)
                        list_total.append(total)
                        list_p99.append(p99)
                        list_avg_latency.append(avg_latency)

                        list_time_begin.append(time_begin)
                        list_time_end.append(time_end)

                        list_bandwidth.append(bandwidth)

                        list_emmc_read_nums.append(emmc_read_nums)
                        list_emmc_read_avg_latency.append(emmc_read_avg_latency)
                        list_emmc_read_p99.append(emmc_read_p99)
                        list_emmc_write_nums.append(emmc_write_nums)
                        list_emmc_write_avg_latency.append(emmc_write_avg_latency)
                        list_emmc_write_p99.append(emmc_write_p99)
                        list_sd_read_nums.append(sd_read_nums)
                        list_sd_read_avg_latency.append(sd_read_avg_latency)
                        list_sd_read_p99.append(sd_read_p99)
                        list_sd_write_nums.append(sd_write_nums)
                        list_sd_write_avg_latency.append(sd_write_avg_latency)
                        list_sd_write_p99.append(sd_write_p99)

    
    list_emmc_kb_read = [x / 1024 for x in list_emmc_kb_read]
    list_emmc_kb_wrtn = [x / 1024 for x in list_emmc_kb_wrtn]
    list_sd_kb_read = [x / 1024 for x in list_sd_kb_read]
    list_sd_kb_wrtn = [x / 1024 for x in list_sd_kb_wrtn]
    dict_cache_policy={'fifo':'FIFO','lfu':"LFU",'lru':'LRU',
                        'lirs':'LIRS','arc':'ARC','clockpro':'CLOCK-Pro',
                        'random':'Random','2q':'2Q','tinylfu':'TinyLFU'}
    list_policy = [dict_cache_policy[x] for x in list_policy]

    # 将多个列表放入一个字典
    data = {'Trace Pattern': list_trace_list, 'Operation Read Ratio': list_operation_read_ratio,
            'IO': list_io, 'Cache Size': list_cache_size, 'Cache Policy': list_policy,
            'Hit Ratio':list_trace_hit_ratio,
            'Average Latency(ms)': list_avg_latency, 'P99 Latency(ms)': list_p99, 
            'Average CPU Usage(%)': list_cpu_usage, 'Average Memory Used(MB)': list_mem_used,
            'Total Time(s)': list_total, "Bandwidth(MB/s)": list_bandwidth,
            'eMMC Read Mount(MB)': list_emmc_kb_read, 'eMMC Write Mount(MB)': list_emmc_kb_wrtn,
            'SD Read Mount(MB)': list_sd_kb_read, 'SD Write Mount(MB)': list_sd_kb_wrtn,
            'eMMC Read Numbers': list_emmc_read_nums, 'eMMC Read Avarage Latency(ms)': list_emmc_read_avg_latency, 'eMMC Read P99 Latency(ms)':list_emmc_read_p99,
            'eMMC Write Numbers': list_emmc_write_nums, 'eMMC Write Avarage Latency(ms)': list_emmc_write_avg_latency, 'eMMC Write P99 Latency(ms)':list_emmc_write_p99,
            'SD Read Numbers': list_sd_read_nums, 'SD Read Avarage Latency(ms)': list_sd_read_avg_latency, 'SD Read P99 Latency(ms)':list_sd_read_p99,
            'SD Write Numbers': list_sd_write_nums, 'SD Write Avarage Latency(ms)': list_sd_write_avg_latency, 'SD Write P99 Latency(ms)':list_sd_write_p99}

    # print(data)

    # 将字典转换为DataFrame
    df = pd.DataFrame(data)

    # 指定要保存的Excel文件名
    excel_file = path_head + 'statistic.xlsx'

    # save_dataframe_to_excel_with_format(df,excel_file)
    # 将DataFrame保存为Excel文件
    df.to_excel(excel_file, index=False)

    print(f"data save in {excel_file}")

# 合并所有statistic.xlsx，存放到path_head目录下
# 输入文件头 path_head 和 存放各个statistic.xlsx的文件list folder_list
def merge_excel_files(path_dir, folder_list):
    # 初始化空 DataFrame 用于合并数据
    combined_df = pd.DataFrame()

    # 遍历文件夹并读取每个文件
    for folder in folder_list:
        file_path = os.path.join(path_dir + folder, 'statistic.xlsx')
        print(file_path)
        if os.path.exists(file_path):
            df = pd.read_excel(file_path)
            combined_df = combined_df.append(df, ignore_index=True)

    # 保存合并后的 DataFrame 到新文件
    combined_df.to_excel(path_dir + 'statistic.xlsx', index=False)
    print("Files merged successfully into",path_dir + 'statistic.xlsx')


def get_trace_type_and_read_ratio_folder():
    """
    获取指定目录下的文件夹名称，并进一步获取每个文件夹下的子文件夹名称。

    参数：
    - path_head：要遍历的目录路径
    """
    global trace_type_list
    global operation_read_ratio_list

    trace_type_list = []
    operation_read_ratio_list = []

    # 获取path_head目录下的所有文件夹
    for item in os.listdir(path_head):
        # 构建完整路径
        item_path = os.path.join(path_head, item)

        # 检查是否为文件夹
        if os.path.isdir(item_path):
            # 将文件夹名称存入trace_type_list
            if item != 'log':
                trace_type_list.append(item)

            # 获取子文件夹名称并存入operation_read_ratio_list
            subfolders = [subfolder for subfolder in os.listdir(item_path) if
                          os.path.isdir(os.path.join(item_path, subfolder))]
            operation_read_ratio_list.extend(subfolders)


def get_folders(directory_path):
    """
    获取指定目录下的所有文件夹名称。

    参数：
    - directory_path：要遍历的目录路径

    返回值：
    - folder_names：所有文件夹的名称列表
    """
    folder_names = []

    items = os.listdir(directory_path)

    for item in items:
        item_path = os.path.join(directory_path, item)

        if os.path.isdir(item_path):
            folder_names.append(item)

    return folder_names


def resetPath():
    global log_dir
    global file_cpu_usage_path
    global file_mem_used_path
    global file_disk_path
    global file_power_path

    log_dir = path_head + 'log/'

    file_cpu_usage_path = log_dir + file_cpu_usage_name
    file_mem_used_path = log_dir + file_mem_used_name
    file_disk_path = log_dir + file_disk_name
    file_power_path = log_dir + file_power_name

    os.chmod(file_cpu_usage_path, 0o666)
    os.chmod(file_mem_used_path, 0o666)
    os.chmod(file_disk_path, 0o666)
    os.chmod(file_power_path, 0o666)

def resetPath_except_power():
    global log_dir
    global file_cpu_usage_path
    global file_mem_used_path
    global file_disk_path

    log_dir = path_head + 'log/'

    file_cpu_usage_path = log_dir + file_cpu_usage_name
    file_mem_used_path = log_dir + file_mem_used_name
    file_disk_path = log_dir + file_disk_name

    os.chmod(file_cpu_usage_path, 0o666)
    os.chmod(file_mem_used_path, 0o666)
    os.chmod(file_disk_path, 0o666)

def run(test_files_path):
    global  path_head
    path_head_folders = get_folders(test_files_path)
    # print(path_head_folders)
    for folder in path_head_folders:
        path_head = test_files_path + folder + '/'
        resetPath()
        get_trace_type_and_read_ratio_folder()
        process_results()

def run_except_power(test_files_path):
    global  path_head
    path_head_folders = get_folders(test_files_path)
    # print(path_head_folders)
    for folder in path_head_folders:
        path_head = test_files_path + folder + '/'
        resetPath_except_power()
        get_trace_type_and_read_ratio_folder()
        process_results_except_power()

# def get_device_id_and_chunk_size_list():
#     global device_id_list
#     global chunk_size_list

#     device_id_list = []
#     chunk_size_list = []

#     for item in os.listdir(path_head):
#         item_path = os.path.join(path_head, item)

#         if os.path.isdir(item_path):
#             if item != 'log':
#                 device_id_list.append(item)

#             sub_folders = [subfolder for subfolder in os.listdir(item_path) if
#                           os.path.isdir(os.path.join(item_path, subfolder))]
#             chunk_size_list.extend(sub_folders)

#     print("device_id_list:",device_id_list)
#     print("chunk_size_list:",chunk_size_list)

def get_device_id_and_chunk_size_list():
    global device_id_list
    global chunk_size_list

    device_id_list = []
    chunk_size_set = set()  # Use a set to avoid duplicate entries

    for item in os.listdir(path_head):
        item_path = os.path.join(path_head, item)

        if os.path.isdir(item_path) and item != 'log':
            device_id_list.append(item)

            sub_folders = [subfolder for subfolder in os.listdir(item_path)
                           if os.path.isdir(os.path.join(item_path, subfolder))]
            chunk_size_set.update(sub_folders)  # Add new items to the set

    chunk_size_list = list(chunk_size_set)  # Convert set back to list if necessary
    chunk_size_list.sort()  # Optional: sort the list if you need ordered output

    print("device_id_list:", device_id_list)
    print("chunk_size_list:", chunk_size_list)


def process_results_for_device_test():
    global device_id_list
    global chunk_size_list

    list_device_id, list_chunk_size = [], []
    list_total, list_p99, list_avg_latency =  [], [], []
    list_cpu_usage, list_mem_used, list_emmc_kb_read, list_emmc_kb_wrtn, list_sd_kb_read, list_sd_kb_wrtn = [], [], [], [], [], []
    list_time_begin, list_time_end = [], []
    list_bandwidth = []
    list_emmc_read_nums, list_emmc_read_avg_latency, list_emmc_read_p99 = [], [], []
    list_emmc_write_nums, list_emmc_write_avg_latency, list_emmc_write_p99 = [], [], []
    list_sd_read_nums, list_sd_read_avg_latency, list_sd_read_p99 = [], [], []
    list_sd_write_nums, list_sd_write_avg_latency, list_sd_write_p99 = [], [], []
    
    rdwr_only = True # ratio == 'read_1' or ratio == 'read_0'
    
    for device_id in device_id_list:
        for chunk_size in chunk_size_list:
            list_device_id.append(device_id)
            list_chunk_size.append(chunk_size)
            file_path_begin = os.path.join(path_head, device_id, chunk_size)
            print('data process:', file_path_begin)
            
            file_statistic_path = os.path.join(file_path_begin, file_statistic_name)

            total, p99, avg_latency, \
            time_begin, time_end, bandwidth, \
            emmc_read_nums, emmc_read_avg_latency, emmc_read_p99, \
            emmc_write_nums, emmc_write_avg_latency, emmc_write_p99, \
            sd_read_nums, sd_read_avg_latency, sd_read_p99, \
            sd_write_nums, sd_write_avg_latency, sd_write_p99 = \
            extract_statistic_device_test(
                file_statistic_path,
                rdwr_only)
            
            list_cpu_usage.append(
                calculate_avg_cpu_usage(file_cpu_usage_path, time_begin,
                                        time_end - timedelta(seconds=1)))
            list_mem_used.append(calculate_avg_mem_used(file_mem_used_path, time_begin, time_end))
            emmc_kb_read, emmc_kb_wrtn, sd_kb_read, sd_kb_wrtn \
                = calculate_disk_read_wrtn(file_disk_path, time_begin, time_end)

            list_emmc_kb_read.append(emmc_kb_read)
            list_emmc_kb_wrtn.append(emmc_kb_wrtn)
            list_sd_kb_read.append(sd_kb_read)
            list_sd_kb_wrtn.append(sd_kb_wrtn)

            list_total.append(total)
            list_p99.append(p99)
            list_avg_latency.append(avg_latency)

            list_time_begin.append(time_begin)
            list_time_end.append(time_end)

            list_bandwidth.append(bandwidth)

            list_emmc_read_nums.append(emmc_read_nums)
            list_emmc_read_avg_latency.append(emmc_read_avg_latency)
            list_emmc_read_p99.append(emmc_read_p99)
            list_emmc_write_nums.append(emmc_write_nums)
            list_emmc_write_avg_latency.append(emmc_write_avg_latency)
            list_emmc_write_p99.append(emmc_write_p99)
            list_sd_read_nums.append(sd_read_nums)
            list_sd_read_avg_latency.append(sd_read_avg_latency)
            list_sd_read_p99.append(sd_read_p99)
            list_sd_write_nums.append(sd_write_nums)
            list_sd_write_avg_latency.append(sd_write_avg_latency)
            list_sd_write_p99.append(sd_write_p99)

    
    list_emmc_kb_read = [x / 1024 for x in list_emmc_kb_read]
    list_emmc_kb_wrtn = [x / 1024 for x in list_emmc_kb_wrtn]
    list_sd_kb_read = [x / 1024 for x in list_sd_kb_read]
    list_sd_kb_wrtn = [x / 1024 for x in list_sd_kb_wrtn]
    dict_device_id={'sd':'SD','emmc':"eMMC"}
    list_device_id = [dict_device_id[x] for x in list_device_id]

    data = {'Device ID': list_device_id, 'Chunk Size(KB)': list_chunk_size,
            'Average Latency(ms)': list_avg_latency, 'P99 Latency(ms)': list_p99, 
            'Average CPU Usage(%)': list_cpu_usage, 'Average Memory Used(MB)': list_mem_used,
            'Total Time(s)': list_total, "Bandwidth(MB/s)": list_bandwidth,}

    # print(data)

    # 将字典转换为DataFrame
    df = pd.DataFrame(data)

    # 指定要保存的Excel文件名
    excel_file = path_head + 'statistic.xlsx'

    # save_dataframe_to_excel_with_format(df,excel_file)
    # 将DataFrame保存为Excel文件
    df.to_excel(excel_file, index=False)

    print(f"data save in {excel_file}")

def run_device_test():
    resetPath_except_power()
    get_device_id_and_chunk_size_list()
    process_results_for_device_test()

file_cpu_usage_name = 'cpu_usage.log'
file_mem_used_name = 'mem_used.log'
file_disk_name = 'disk_read_wrtn.log'
file_power_name = 'e-test.xlsx'
file_statistic_name = 'statistic.txt'

io_list = ['io_off', 'io_on']
cache_size_list = ['0.02', '0.04', '0.06', '0.08', '0.1']
policy_list = ['fifo', 'lru', 'lfu', 'lirs', 'arc', 'clockpro', 'random', '2q', 'tinylfu']

trace_type_list = []
operation_read_ratio_list = []

# for run_device
device_id_list = []
chunk_size_list = []

path_head = ''
log_dir = ''
file_cpu_usage_path = ''
file_mem_used_path = ''
file_disk_path = ''
file_power_path = ''

if __name__ == '__main__':
    # path_dir = 'E:/projects/Caching-Policy/records/'
    # folder_list = ['latest']
    # # folder_list = ['zipfian','latest','uniform', ]
    # for folder in folder_list:
    #     run_except_power(path_dir+folder+'/')

    # for folder in folder_list:
    #     sub_folder_head = path_dir+folder+'/'
    #     sub_folder_list = get_folders(sub_folder_head)
    #     merge_excel_files(sub_folder_head, sub_folder_list)
    # merge_excel_files(path_dir,folder_list)
    path_head = 'E:/records/2024-05-03_11-38-30_uniform_read_1_5GB/'
    run_device_test()
